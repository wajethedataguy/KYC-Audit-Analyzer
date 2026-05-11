# ─────────────────────────────────────────────────────────────
# ui.py — AAU KYC Reports Analyzer (Desktop App)
# ─────────────────────────────────────────────────────────────

import os
import re
from dateutil.parser import parse
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog

import pandas as pd
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# 🧹 Internal Modules
from .file_loader import load_file
from .processor import clean_dataframe
from .utils import get_filename, logic_registry
from .logic_loader import load_all_logics
from .merge_engine import process_uploaded_files
from KYC_Viewer.logic_metadata import logic_metadata
from KYC_Viewer.utils import logic_registry


KYC_NEW_ACCOUNTS_COLS = [
    "Serial_No",
    "rm_code",
    "branch_code",
    "customer_id",
    "accountno",
    "account_type_(new)",
    "iban",
    "account_title",
    "currency",
    "category",
    "category_desc",
    "working_balance",
    "actual_currency",
    "limit_ref",
    "ac_open_date",
    "posting_rest",
    "zakat_exempt",
    "fiqah",
    "referee",
    "summary_date",
]

KYC_OLD_ACCOUNTS_HT_COLS = [
    "COB_DATE",
    "Branch",
    "Br. Code & Name",
    "Customer ID",
    "Account No.",
    "Account Title",
    "Account Type",
    "Entity Type (Sector)",
    "Currency",
    "Purpose of Account",
    "Occupation",
    "Nature of Business",
    "Account T/O (P.A KYC)",
    "Account T/O With Tolerence",
    "Account T/O With Tolerence Neg",
    "Total_Actual_TO",
    "ACTUAL_DEBIT_TO",
    "ACTUAL_CREDIT_TO",
    "ID No. of Customer",
    "CategoryType",
    "Company",
]

KYC_OLD_ACCOUNTS_INDIV_COLS = [
    "Account_numumber",
    "Customer_Num",
    "Name",
    "Account_Open_Dt",
    "Branch_Code",
    "Branch_name",
    "JOINT_HOLDER",
    "ACCT_OPR_INST_1",
    "ACCT_OPR_INST_2",
    "Annual_TO_GT_M",
    "Annual_TO_GT_M1",
    "EXP_MONTH_TOVER",
    "KYC_ATO",
    "KYC_CO_ATO",
    "KYC_REASON_HIGH",
    "MODEDEPOSITS",
    "MODEWITHDRAW",
    "MON_TOVER_CORP",
    "MON_TOVER_CRG",
    "MONTH_TOVER_RG",
    "NTN_FILER",
    "Purpose",
    "Source_Funds_Other",
    "UNSCLISTST",
    "FIQAH",
    "ZAKAT_EXEMPT",
    "CURRENCY",
    "INACTIV_MARKER",
    "CATEGORY",
    "Product_Desc",
    "Status_of_Account",
    "Ac_turnover_kyc",
    "Risk_Level",
    "CS_POS",
    "SOURCE_OF_INCOME",
    "OCCUPATION",
    "NAME_OF_EMP",
    "CURRENT_SALARY",
    "NAME_OF_BUS",
    "REL_POILITICAL",
    "STATUS",
    "KYC_REVW_COMENT",
]


# 🔄 Load all logic modules into registry
load_all_logics()


class Tooltip:
    """Creates a tooltip for a given widget"""

    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)
        self.selected_option = tk.StringVar()

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")  # get widget bounding box
        x += self.widget.winfo_rootx() + 20
        y += self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # Remove window decorations
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=self.text,
            justify="left",
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            font=("Segoe UI", 9),
        )
        label.pack(ipadx=4, ipady=2)

    def hide_tip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


# ─────────────────────────────────────────────────────────────
# Main Viewer Class
# ─────────────────────────────────────────────────────────────
class ExcelCSVViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("AAU Team")
        self.root.geometry("1000x600")
        self.root.configure(bg="#f0f4f7")
        self.detail_tree = None
        self.detail_title_label = None

        # 🔧 Internal State
        self.uploaded_paths = []
        self.files_data = {}
        self.current_file = tk.StringVar()
        self.current_sheet = tk.StringVar()
        self.filtered_df = None
        self.is_dark_mode = False
        self.selected_logic = tk.StringVar()
        self.kyc_new_final = None
        self.kyc_ht_final = None
        self.kyc_ind_final = None
        # ✅ Populate logic names AFTER loading all logic modules
        self.logic_names = sorted(list(logic_registry.keys()))
        self.selected_logic.set(self.logic_names[0] if self.logic_names else "")

        # 🔢 Observation grid state
        self.selected_observation = tk.IntVar(value=1)
        self.obs_buttons = {}

        # 🎨 UI Setup
        self.setup_styles()
        self.create_widgets()

    # ─────────────────────────────────────────────────────────
    # UI Styling
    # ─────────────────────────────────────────────────────────
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=6)
        style.configure("TLabel", font=("Segoe UI", 10), background="#f0f4f7")
        style.configure("TCombobox", padding=4)

    # ─────────────────────────────────────────────────────────
    # UI Layout and Widgets
    # ─────────────────────────────────────────────────────────
    def create_widgets(self):
        # Root window background
        self.root.configure(background="white")

        # Layout tuning
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(3, weight=1)  # content frame will expand
        self.root.rowconfigure(7, weight=0)

        # ------------------- Header -------------------
        header_frame = tk.Frame(self.root, bg="#DA1E26", height=64)
        header_frame.grid(row=0, column=0, sticky="ew")
        header_frame.grid_propagate(False)

        title_label = tk.Label(
            header_frame,
            text="Audit & Inspection Group - KYC Exceptions Generator",
            font=("Arial Black", 18, "bold"),
            fg="white",
            bg="#DA1E26",
        )
        title_label.pack(side="left", padx=20, pady=8)

        # ------------------- Upload Section -------------------
        upload_frame = tk.LabelFrame(
            self.root, text="📁 File Upload", padx=8, pady=8, bg="white", fg="#000000"
        )
        upload_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(12, 6))

        # 🔑 Use 18 columns for PERFECT alignment
        for i in range(18):
            upload_frame.columnconfigure(i, weight=1)

        btn_font = ("Arial Black", 11, "bold")
        btn_bg = "#DA1E26"
        btn_fg = "white"

        upload_btn = tk.Button(
            upload_frame,
            text="Upload Excel/CSV Files",
            bg=btn_bg,
            fg=btn_fg,
            font=btn_font,
            command=self.upload_files,
        )
        upload_btn.grid(row=0, column=0, columnspan=6, sticky="ew", padx=8, pady=6)

        gen_btn = tk.Button(
            upload_frame,
            text="⚙️ Generate Individual Exception Table",
            bg=btn_bg,
            fg=btn_fg,
            font=btn_font,
            command=lambda: self.run_logic_and_save(),  # new wrapper
        )

        gen_btn.grid(row=0, column=6, columnspan=6, sticky="ew", padx=8, pady=6)

        excel_btn = tk.Button(
            upload_frame,
            text="📊 Generate Excel Report",
            bg=btn_bg,
            fg=btn_fg,
            font=btn_font,
            command=self.generate_excel_workbook_from_ui,
        )
        excel_btn.grid(row=0, column=12, columnspan=6, sticky="ew", padx=8, pady=6)

        # ------------------ Unified Upload + Sampling Section ------------------
        section_frame = tk.Frame(self.root, bg="white")
        section_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(12, 6))

        # 🔑 Define a consistent 18-column grid for the whole section
        for i in range(18):
            section_frame.columnconfigure(i, weight=1)

        # ------------------- Upload Row -------------------
        upload_btn = tk.Button(
            section_frame,
            text="Upload Excel/CSV Files",
            bg=btn_bg,
            fg=btn_fg,
            font=btn_font,
            command=self.upload_files,
        )
        upload_btn.grid(row=0, column=0, columnspan=6, sticky="ew", padx=8, pady=6)

        gen_btn = tk.Button(
            section_frame,
            text="⚙️ Generate Individual Exception Table",
            bg=btn_bg,
            fg=btn_fg,
            font=btn_font,
            command=lambda: self.run_logic_and_save(),
        )
        gen_btn.grid(row=0, column=6, columnspan=6, sticky="ew", padx=8, pady=6)

        excel_btn = tk.Button(
            section_frame,
            text="📊 Generate Excel Report",
            bg=btn_bg,
            fg=btn_fg,
            font=btn_font,
            command=self.generate_excel_workbook_from_ui,
        )
        excel_btn.grid(row=0, column=12, columnspan=6, sticky="ew", padx=8, pady=6)

        # ------------------ Sampling + Audit Period Row ------------------
        self.selected_option = tk.StringVar(value="Choose Filtering Type")
        sampling_btn = tk.Button(
            section_frame,
            text="Select Filter",
            bg=btn_bg,
            fg=btn_fg,
            font=("Arial Black", 10, "bold"),
            command=self.handle_program_sampling,
        )
        sampling_btn.grid(row=1, column=0, columnspan=3, padx=8, pady=6, sticky="ew")

        sampling_dropdown = ttk.Combobox(
            section_frame,
            textvariable=self.selected_option,
            values=["Program Based Instances", "Random Instances"],
            state="readonly",
        )
        sampling_dropdown.grid(
            row=1, column=3, columnspan=3, padx=8, pady=6, sticky="ew"
        )

        # Start/End Date (middle group)
        start_label = tk.Label(
            section_frame, text="Start Date:", bg="white", font=("Arial Black", 10)
        )
        start_label.grid(row=1, column=6, sticky="e", padx=6, pady=6)

        self.start_entry = tk.Entry(section_frame, bg="#DCDCDC", font=("Segoe UI", 10))
        self.start_entry.grid(
            row=1, column=7, columnspan=2, padx=6, pady=6, sticky="ew"
        )
        self.add_placeholder(self.start_entry, "YYYYMMDD")

        end_label = tk.Label(
            section_frame, text="End Date:", bg="white", font=("Arial Black", 10)
        )
        end_label.grid(row=1, column=9, sticky="e", padx=6, pady=6)

        self.end_entry = tk.Entry(section_frame, bg="#DCDCDC", font=("Segoe UI", 10))
        self.end_entry.grid(row=1, column=10, columnspan=2, padx=6, pady=6, sticky="ew")
        self.add_placeholder(self.end_entry, "YYYYMMDD")

        # Right group frame spanning same 6 columns as excel_btn
        right_frame = tk.Frame(section_frame, bg="white")
        right_frame.grid(row=1, column=12, columnspan=6, sticky="ew", padx=8, pady=6)

        # Make two equal columns inside
        right_frame.columnconfigure(0, weight=1)
        right_frame.columnconfigure(1, weight=1)

        # Audit Period Button
        audit_btn = tk.Button(
            right_frame,
            text="Add Audit Period",
            bg=btn_bg,
            fg=btn_fg,
            font=("Arial Black", 10, "bold"),
            command=self.apply_audit_period,
        )
        audit_btn.grid(row=0, column=0, padx=4, pady=6, sticky="ew")

        # Refresh Button
        refresh_btn = tk.Button(
            right_frame,
            text="🔄 Refresh",
            bg=btn_bg,
            fg=btn_fg,
            font=("Arial Black", 10, "bold"),
            command=self.refresh_app,
        )
        refresh_btn.grid(row=0, column=1, padx=4, pady=6, sticky="ew")

        # ------------------- Content Frame -------------------
        content_frame = tk.Frame(self.root, bg="white")
        content_frame.grid(row=3, column=0, sticky="nsew", padx=20, pady=10)
        content_frame.columnconfigure(0, weight=1)
        content_frame.rowconfigure(1, weight=1)

        # Branch frame
        branch_frame = tk.LabelFrame(
            content_frame,
            text="Branch Details",
            padx=10,
            pady=10,
            bg="white",
            fg="#000000",
        )
        branch_frame.grid(row=0, column=0, sticky="ew")
        self.branch_label = tk.Label(
            branch_frame,
            text="Branch: [Not determined yet]",
            font=("Segoe UI", 11, "bold"),
            bg="white",
            fg="#000000",
        )
        self.branch_label.grid(row=0, column=0, sticky="w")

        # ------------------- Main Split -------------------
        main_frame = tk.Frame(content_frame, bg="white")
        main_frame.grid(row=1, column=0, sticky="nsew", pady=(10, 0))
        main_frame.columnconfigure(0, weight=0)  # left fixed
        main_frame.columnconfigure(1, weight=1)  # right expands
        main_frame.rowconfigure(0, weight=1)

        # ------------------- Left: Exceptions -------------------
        obs_frame = tk.LabelFrame(
            main_frame,
            text="Exceptions",
            padx=5,
            pady=5,
            bg="white",
            fg="#000000",
            width=250,
            height=400,
        )
        obs_frame.grid(row=0, column=0, sticky="ns", padx=(0, 10))
        obs_frame.grid_propagate(False)

        # ✅ Add weight so canvas expands
        obs_frame.columnconfigure(0, weight=1)
        obs_frame.rowconfigure(0, weight=1)

        # Canvas for scrollable content
        obs_canvas = tk.Canvas(
            obs_frame, borderwidth=0, highlightthickness=0, bg="white"
        )
        obs_canvas.grid(row=0, column=0, sticky="nsew")

        # Vertical scrollbar
        obs_scrollbar = tk.Scrollbar(
            obs_frame, orient="vertical", command=obs_canvas.yview
        )
        obs_scrollbar.grid(row=0, column=1, sticky="ns")
        obs_canvas.configure(yscrollcommand=obs_scrollbar.set)

        # Optional: Horizontal scrollbar if button labels are wide
        obs_scrollbar_x = tk.Scrollbar(
            obs_frame, orient="horizontal", command=obs_canvas.xview
        )
        obs_scrollbar_x.grid(row=1, column=0, sticky="ew")
        obs_canvas.configure(xscrollcommand=obs_scrollbar_x.set)

        # Inner frame for buttons
        obs_inner = tk.Frame(obs_canvas, bg="white")
        obs_canvas.create_window((0, 0), window=obs_inner, anchor="nw")

        # Update scrollregion when inner frame changes
        obs_inner.bind(
            "<Configure>",
            lambda e: obs_canvas.configure(scrollregion=obs_canvas.bbox("all")),
        )

        # Populate with your logic buttons
        self.create_observation_grid(obs_inner)

        # ------------------- Right: Observation Details -------------------
        detail_frame = tk.LabelFrame(
            main_frame,
            text="Observation Details",
            padx=5,
            pady=5,
            bg="white",
            fg="#000000",
        )
        detail_frame.grid(row=0, column=1, sticky="nsew")
        detail_frame.columnconfigure(0, weight=1)
        detail_frame.rowconfigure(1, weight=1)

        self.detail_title_label = tk.Label(
            detail_frame,
            text="Observation 1: [no logic linked]",
            font=("Segoe UI", 10, "bold"),
            bg="white",
            fg="#000000",
        )
        self.detail_title_label.grid(row=0, column=0, sticky="w", padx=5, pady=(0, 5))

        table_frame = tk.Frame(detail_frame, bg="white")
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        self.detail_tree = ttk.Treeview(table_frame, columns=(), show="headings")
        self.detail_tree.grid(row=0, column=0, sticky="nsew")

        # Add vertical scrollbar
        tree_scroll_y = tk.Scrollbar(
            table_frame, orient="vertical", command=self.detail_tree.yview
        )
        tree_scroll_y.grid(row=0, column=1, sticky="ns")

        # Add horizontal scrollbar
        tree_scroll_x = tk.Scrollbar(
            table_frame, orient="horizontal", command=self.detail_tree.xview
        )
        tree_scroll_x.grid(row=1, column=0, sticky="ew")

        self.detail_tree.configure(
            yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set
        )

        # ------------------- Status & Summary -------------------
        status_frame = tk.Frame(self.root, bg="white")
        status_frame.grid(row=7, column=0, sticky="ew", padx=20, pady=(0, 10))

        self.status_label = tk.Label(
            status_frame, text="Status: Waiting", fg="#2980b9", bg="white"
        )
        self.status_label.grid(row=0, column=0, sticky="w")

        self.summary_label = tk.Label(
            status_frame, text="Summary: N/A", fg="#7f8c8d", bg="white"
        )
        self.summary_label.grid(row=0, column=1, sticky="e")

    def run_logic_and_save(self):
        self.report_mode = False
        logic_name = self.selected_logic.get()
        self.run_logic_and_show(logic_name)
        if hasattr(self, "filtered_df") and not self.filtered_df.empty:
            filename = self.prompt_for_filename()
            if filename:
                full_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    initialfile=f"{filename}.xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                )
                if full_path:
                    self.filtered_df.to_excel(full_path, index=False)
                    messagebox.showinfo("Saved", f"Report saved to:\n{full_path}")

    def add_placeholder(self, entry, placeholder):
        entry.insert(0, placeholder)
        entry.placeholder = placeholder
        entry.config(fg="grey")

        def on_focus_in(event):
            if entry.get() == entry.placeholder:
                entry.delete(0, tk.END)
                entry.config(fg="black")

        def on_focus_out(event):
            if entry.get().strip() == "":
                entry.insert(0, entry.placeholder)
                entry.config(fg="grey")

        entry.bind("<FocusIn>", on_focus_in)
        entry.bind("<FocusOut>", on_focus_out)

    # ─────────────────────────────────────────────────────────
    # Refresh Function
    # ─────────────────────────────────────────────────────────
    def refresh_app(self):
        # ⚡ Do NOT clear uploaded files
        # self.uploaded_paths = []
        # self.files_data = {}

        # Reset current selections and state
        self.current_file.set("")
        self.current_sheet.set("")
        self.filtered_df = None
        self.kyc_new_final = None
        self.kyc_ht_final = None
        self.kyc_ind_final = None

        # Reset dropdowns and entries
        self.selected_option.set("Choose Filtering Type")
        self.start_entry.delete(0, tk.END)
        self.end_entry.delete(0, tk.END)
        self.add_placeholder(self.start_entry, "YYYYMMDD")
        self.add_placeholder(self.end_entry, "YYYYMMDD")

        # Reset labels
        self.branch_label.config(text="Branch: [Not determined yet]")
        self.status_label.config(text="Status: Waiting", fg="#2980b9")
        self.summary_label.config(text="Summary: N/A", fg="#7f8c8d")

        # Clear detail tree
        if self.detail_tree is not None:
            self.detail_tree.delete(*self.detail_tree.get_children())

        # Reset observation grid
        self.selected_observation.set(1)
        self.detail_title_label.config(text="Observation 1: [no logic linked]")

        print("✔ Application state refreshed. Uploaded files preserved.")

    # ------------------------------------------------------------
    # Sampling Handler
    # ------------------------------------------------------------
    def on_select_sampling_button_click(self):
        # Run Program Sampling diagnostic and collect distinct accounts
        self.distinct_accounts_for_sampling = self.extract_distinct_accounts()
        print(
            f"Distinct Account Numbers collected: {self.distinct_accounts_for_sampling}"
        )
        messagebox.showinfo(
            "Program Based Instances",
            f"{len(self.distinct_accounts_for_sampling)} distinct accounts collected.",
        )

    # ------------------- Sampling Handler -------------------
    def handle_program_sampling(self):
        sampling_choice = self.selected_option.get()
        if sampling_choice == "Random Instances":
            self.selected_sampling_type = "Random Instances"
            print("Filtering type set to Random Instances.")
            messagebox.showinfo("Random Instances", "Random Instances selected.")
        elif sampling_choice == "Program Based Instances":
            self.selected_sampling_type = "Program Based Instances"
            print("Sampling type set to Program Based Instances.")
            messagebox.showinfo(
                "Program Based Instances", "Program Based Instances selected."
            )
        else:
            messagebox.showwarning("Sampling", "Please select a valid sampling type.")

    # ------------------- Normalization Helper -------------------
    def normalize_account(self, value):
        """Normalize account numbers into consistent 10-digit string form, preserving leading zeros."""
        if pd.isna(value):
            return ""
        try:
            val_str = str(value).strip()
            # keep only digits
            if val_str.isdigit():
                # pad shorter values with leading zeros
                if len(val_str) < 10:
                    val_str = val_str.zfill(10)
                return val_str
            return ""
        except Exception:
            return ""

    # ------------------- apply_program_sampling_to_kyc -------------------
    def apply_program_sampling_to_kyc(self, distinct_accounts: set):
        """
        Match extracted exception accounts with KYC sheets and
        add Program_Sample column containing the matched account number.
        """

        if not distinct_accounts:
            print("⚠️ No distinct accounts provided for Program Based Instances.")
            return

        # Normalize exception accounts once
        normalized_exceptions = {
            self.normalize_account(acc) for acc in distinct_accounts if acc
        }

        def apply_to_df(df, account_col):
            if df is None or df.empty or account_col not in df.columns:
                return df

            # Normalize population accounts
            df[account_col] = df[account_col].apply(self.normalize_account)

            # Create Program_Sample column
            df["Program_Based _Instances"] = df[account_col].apply(
                lambda x: x if x in normalized_exceptions else ""
            )

            matched = (df["Program_Based _Instances"] != "").sum()
            print(
                f"✅ {matched} accounts matched in sheet using column '{account_col}'"
            )
            return df

        # Apply to each sheet with the correct column
        self.kyc_new_final = apply_to_df(self.kyc_new_final, "accountno")
        self.kyc_ht_final = apply_to_df(self.kyc_ht_final, "Account No.")
        self.kyc_ind_final = apply_to_df(self.kyc_ind_final, "Account_numumber")

    # ------------------- Collect all exception records -------------------
    def extract_distinct_accounts(self):
        """
        Collect distinct integer-normalized account numbers directly from logic outputs (exceptions only).
        No dependency on generated Excel file.
        """
        try:
            combined_accounts = set()
            candidate_cols = [
                "Account_Number",
                "ACCOUNT_NUMBER",
                "accountno",
                "Account No.",
                "Account_numumber",
            ]

            for logic_name, info in self.logic_outputs.items():
                df_logic = info.get("df")
                if isinstance(df_logic, pd.DataFrame):
                    # find any matching column
                    for col in candidate_cols:
                        if col in df_logic.columns:
                            accounts = (
                                df_logic[col].dropna().astype(str).str.strip().tolist()
                            )
                            for acc in accounts:
                                norm_acc = self.normalize_account(acc)
                                if norm_acc:
                                    combined_accounts.add(norm_acc)
                            break  # stop after first matching column

            print(f"✔ Total exception accounts collected: {len(combined_accounts)}")
            return combined_accounts

        except Exception as e:
            print("Error extracting accounts:", e)
            return set()

    # ------------------- analyze_account_age -------------------
    def apply_audit_period(self):
        """
        Stores the audit period only.
        No filtering, no classification, no KYC access.
        """
        start_str = self.start_entry.get().strip()
        end_str = self.end_entry.get().strip()

        if not start_str or not end_str:
            messagebox.showerror("Input Error", "Please enter Start and End date.")
            return

        # Validate YYYYMMDD format
        try:
            int(start_str)
            int(end_str)
        except:
            messagebox.showerror("Format Error", "Dates must be YYYYMMDD numbers.")
            return

        # Store safely
        self.audit_start_date = start_str
        self.audit_end_date = end_str

        messagebox.showinfo(
            "Audit Period Saved",
            f"Start: {start_str}\nEnd: {end_str}\n\n"
            "Entered audit period has been saved.",
        )

    def convert_date_to_int(self, date_str):
        try:
            dt = pd.to_datetime(date_str, errors="coerce")
            if pd.isna(dt):
                return None
            return int(dt.strftime("%Y%m%d"))
        except:
            return None

    # ============================================================
    # build_sampling_population
    # ============================================================
    def build_sampling_population(self, dataframes: dict) -> pd.DataFrame:
        """
        Merge merged_file.xlsx (metadata) with turnover file
        for Random Instances.
        """

        df1, df2 = None, None

        # ------------------------------------------------------------
        # 1️⃣ Detect merged_file.xlsx (MAIN FILE)
        # ------------------------------------------------------------
        if "merged_file.xlsx" in dataframes:
            sheets_dict, _ = dataframes["merged_file.xlsx"]

            if not sheets_dict:
                raise ValueError("merged_file.xlsx has no sheets.")

            # If system created it, sheet name is usually "Merged"
            if "Merged" in sheets_dict:
                df1 = sheets_dict["Merged"].copy()
            else:
                df1 = list(sheets_dict.values())[0].copy()

            # Normalize columns
            df1.columns = (
                df1.columns.str.strip()
                .str.upper()
                .str.replace(" ", "_")
                .str.replace("-", "_")
                .str.replace(r"[^\w]", "_", regex=True)
            )

            print("✔ merged_file.xlsx detected as metadata file")
        else:
            raise ValueError("merged_file.xlsx is missing.")

        # ------------------------------------------------------------
        # 2️⃣ Detect Turnover File by Signature Columns
        # ------------------------------------------------------------
        for filename, (sheets_dict, path) in dataframes.items():

            # Skip merged file
            if filename == "merged_file.xlsx":
                continue

            for sheet_name, df in sheets_dict.items():

                if df is None or df.empty:
                    continue

                df_norm = df.copy()
                df_norm.columns = (
                    df_norm.columns.str.strip()
                    .str.upper()
                    .str.replace(" ", "_")
                    .str.replace("-", "_")
                    .str.replace(r"[^\w]", "_", regex=True)
                )

                # Signature columns for turnover file
                signature_cols = {
                    "ACCOUNT_NO_",
                    "TOTAL_ACTUAL_TO",
                    "ACCOUNT_T_O__P_A_KYC_",
                    "ACTUAL_DEBIT_TO",
                    "ACTUAL_CREDIT_TO",
                }

                if signature_cols <= set(df_norm.columns):
                    df2 = df_norm
                    print(f"✔ Turnover file detected: {filename}")
                    break

            if df2 is not None:
                break

        if df2 is None:
            raise ValueError("Turnover file not detected. Required columns missing.")

        # ------------------------------------------------------------
        # 3️⃣ Validate Required Columns in merged_file.xlsx
        # ------------------------------------------------------------
        required_metadata_cols = {
            "ACCOUNT_NUMBER",
            "ACCOUNT_OPEN_DT",
            "CUSTOMER_OCCUPATION",
        }

        if not required_metadata_cols <= set(df1.columns):
            missing = required_metadata_cols - set(df1.columns)
            raise ValueError(f"Missing required metadata columns: {missing}")

        # ------------------------------------------------------------
        # 4️⃣ Normalize Join Keys
        # ------------------------------------------------------------
        df1["ACCOUNT_NUMBER"] = df1["ACCOUNT_NUMBER"].fillna("").astype(str).str.strip()

        df2["ACCOUNT_NO_"] = df2["ACCOUNT_NO_"].fillna("").astype(str).str.strip()

        # ------------------------------------------------------------
        # 5️⃣ Merge
        # ------------------------------------------------------------
        merged = pd.merge(
            df1,
            df2[["ACCOUNT_NO_", "TOTAL_ACTUAL_TO"]],
            left_on="ACCOUNT_NUMBER",
            right_on="ACCOUNT_NO_",
            how="left",
        )

        print(f"✔ Sampling population built. Rows: {len(merged)}")

        return merged

    # ============================================================
    # NEW ACCOUNTS SAMPLING WITH DIAGNOSTICS
    # ============================================================
    def approved_sampling(self, df):
        if df is None or df.empty:
            print("⚠️ approved_sampling called with empty or None DataFrame")
            return pd.DataFrame()

        # --- Step 1: detect account number column dynamically ---
        candidate_cols = [
            "ACCOUNT_NUMBER",
            "accountno",
            "Account No.",
            "Account_numumber",
        ]
        account_col = next((c for c in candidate_cols if c in df.columns), None)
        if account_col is None:
            raise ValueError(
                "No account number column found in dataframe for Approved Sampling"
            )

        # Normalize account numbers
        df = df.copy()
        df.loc[:, "account_no_norm"] = df[account_col].apply(self.normalize_account)

        # --- Step 2: filter by exception accounts ---
        exception_accounts = self.extract_distinct_accounts()
        df = df[df["account_no_norm"].isin(exception_accounts)].copy()

        print(f"✔ Total exception accounts collected: {len(exception_accounts)}")
        print(f"✔ Exception accounts matched in merged population: {len(df)}")

        # Optional: show unmatched accounts for debugging
        unmatched = exception_accounts - set(df["account_no_norm"])
        print(f"⚠️ Unmatched accounts: {len(unmatched)}")
        if unmatched:
            print("Sample unmatched accounts:", list(unmatched)[:5])

        # --- Step 3: apply audit period ---
        if hasattr(self, "audit_start_date"):
            start_int = int(self.audit_start_date)

            def clean_date(v):
                if pd.isna(v):
                    return None
                s = str(v).strip()
                s = re.sub(r"\s+\d{2}:\d{2}:\d{2}", "", s)
                try:
                    d = parse(s, dayfirst=True)
                    return int(d.strftime("%Y%m%d"))
                except Exception:
                    return None

            if "ACCOUNT_OPEN_DT" in df.columns:
                df.loc[:, "open_date_int"] = df["ACCOUNT_OPEN_DT"].apply(clean_date)

                # classify accounts
                df.loc[:, "is_new_account"] = df["open_date_int"].apply(
                    lambda x: True if x is not None and x >= start_int else False
                )

        # --- Step 4: ensure 'is_new_account' column exists ---
        if "is_new_account" not in df.columns:
            print(
                "⚠️ 'is_new_account' column missing, adding default False for all rows"
            )
            df.loc[:, "is_new_account"] = False  # treat all as old if missing

        df_new = df[df["is_new_account"] == True].copy()
        df_old = df[df["is_new_account"] == False].copy()

        print("===== ACCOUNT AGE ANALYSIS =====")
        print(f"Total: {len(df)}")
        print(f"Old accounts: {len(df_old)}")
        print(f"New accounts: {len(df_new)}")

        # --- Helper: refill to enforce quotas ---
        def refill_group(candidates, current_sample, target, label):
            taken_ids = (
                set(current_sample["account_no_norm"])
                if "account_no_norm" in current_sample
                else set()
            )
            available = candidates[~candidates["account_no_norm"].isin(taken_ids)]
            needed = target - len(current_sample)
            if needed > 0 and not available.empty:
                extra = available.sample(n=min(needed, len(available)), random_state=42)
                print(f"Refilled {len(extra)} accounts for {label}")
                return pd.concat([current_sample, extra], ignore_index=True)
            return current_sample

        # --- New accounts sampling (dynamic quota based on audit criteria) ---
        new_sample = pd.DataFrame()
        new_count = len(df_new)

        if new_count < 500:
            new_quota = min(50, new_count // 2)  # 50% or 50 whichever is lower
        elif 500 <= new_count < 750:
            new_quota = min(75, new_count)
        else:  # new_count >= 750
            new_quota = min(100, new_count)

        if new_count > 0:
            new_sample = df_new.sample(n=new_quota, random_state=42)
        new_sample = refill_group(df_new, new_sample, new_quota, "New Accounts")

        # --- Old accounts sampling ---
        old_sample_high = pd.DataFrame()
        old_sample_other = pd.DataFrame()

        # 1️⃣ Top 30 high turnover
        if "TOTAL_ACTUAL_TO" in df_old.columns and len(df_old) > 0:
            df_old = df_old.copy()
            # Convert fully to numeric, coerce errors to NaN, drop NaNs
            df_old.loc[:, "TOTAL_ACTUAL_TO_NUM"] = pd.to_numeric(
                df_old["TOTAL_ACTUAL_TO"].astype(str).str.replace(",", "").str.strip(),
                errors="coerce",
            )
            df_old_valid = df_old[df_old["TOTAL_ACTUAL_TO_NUM"].notna()]
            n_high = min(30, len(df_old_valid))
            if n_high > 0:
                old_sample_high = df_old_valid.nlargest(n_high, "TOTAL_ACTUAL_TO_NUM")

        # 2️⃣ Other categories: distribute quota across occupation groups
        if "CUSTOMER_OCCUPATION" in df_old.columns:
            df_old.loc[:, "CUSTOMER_OCCUPATION"] = (
                df_old["CUSTOMER_OCCUPATION"]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.lower()
            )

            categories = [
                "business",
                "salaried",
                "landlord",
                "others",
                "student",
                "house wife",
                "retired",
                "minor",
            ]

            # Exclude accounts already selected in high turnover
            taken_ids = (
                set(old_sample_high["account_no_norm"])
                if "account_no_norm" in old_sample_high
                else set()
            )
            df_occ = df_old[df_old["CUSTOMER_OCCUPATION"].isin(categories)].copy()
            df_occ = df_occ[~df_occ["account_no_norm"].isin(taken_ids)]

            if not df_occ.empty:
                available_categories = df_occ["CUSTOMER_OCCUPATION"].unique().tolist()
                num_cats = len(available_categories)

                # Base quota per category
                base_quota = 30 // num_cats
                remainder = 30 % num_cats

                sampled_rows = []
                shortfall = 0

                # First pass: sample base_quota from each category
                for cat in available_categories:
                    group = df_occ[df_occ["CUSTOMER_OCCUPATION"] == cat]
                    n = min(len(group), base_quota)
                    if n > 0:
                        sampled_rows.append(group.sample(n=n, random_state=42))
                    shortfall += (base_quota - n) if len(group) < base_quota else 0

                # Second pass: distribute remainder + shortfall across categories with extra capacity
                remaining_needed = remainder + shortfall
                if remaining_needed > 0:
                    for cat in available_categories:
                        if remaining_needed <= 0:
                            break
                        group = df_occ[df_occ["CUSTOMER_OCCUPATION"] == cat]
                        already_taken = (
                            pd.concat(sampled_rows) if sampled_rows else pd.DataFrame()
                        )
                        taken_ids = (
                            set(already_taken["account_no_norm"])
                            if "account_no_norm" in already_taken
                            else set()
                        )
                        available = group[~group["account_no_norm"].isin(taken_ids)]
                        if not available.empty:
                            extra_n = min(len(available), remaining_needed)
                            sampled_rows.append(
                                available.sample(n=extra_n, random_state=42)
                            )
                            remaining_needed -= extra_n

                # Final combined sample
                old_sample_other = (
                    pd.concat(sampled_rows, ignore_index=True)
                    if sampled_rows
                    else pd.DataFrame()
                )

                # Ensure exactly 30 accounts
                if len(old_sample_other) > 30:
                    old_sample_other = old_sample_other.sample(n=30, random_state=42)

                # --- Apply refill to enforce quotas ---
                old_sample_high = refill_group(
                    df_old_valid, old_sample_high, 30, "High Turnover"
                )
                old_sample_other = refill_group(
                    df_occ, old_sample_other, 30, "Occupation"
                )

                # --- Detect overlaps across groups and refill quotas ---
                # --- Final overlap-aware refill ---
                def enforce_quota(
                    candidates, current_sample, target, label, other_samples
                ):
                    taken_ids = (
                        set(pd.concat(other_samples)["account_no_norm"])
                        if other_samples
                        else set()
                    )
                    available = candidates[
                        ~candidates["account_no_norm"].isin(taken_ids)
                    ]
                    needed = target - len(current_sample)
                    if needed > 0 and not available.empty:
                        extra = available.sample(
                            n=min(needed, len(available)), random_state=42
                        )
                        print(f"Final refill {len(extra)} accounts for {label}")
                        return pd.concat([current_sample, extra], ignore_index=True)
                    return current_sample

                new_sample = enforce_quota(
                    df_new,
                    new_sample,
                    50,
                    "New Accounts",
                    [old_sample_high, old_sample_other],
                )
                old_sample_high = enforce_quota(
                    df_old_valid,
                    old_sample_high,
                    30,
                    "High Turnover",
                    [new_sample, old_sample_other],
                )
                old_sample_other = enforce_quota(
                    df_occ,
                    old_sample_other,
                    30,
                    "Occupation",
                    [new_sample, old_sample_high],
                )

                # Recombine after final refill
                all_sampled = pd.concat(
                    [new_sample, old_sample_high, old_sample_other], ignore_index=True
                )
                sampled_accounts = set(all_sampled["account_no_norm"].tolist())

        # Combine old samples
        old_sample = pd.concat([old_sample_high, old_sample_other])
        sampled_accounts.update(old_sample["account_no_norm"].unique().tolist())

        # --- Step 5: mark Approved_Sample column ---
        df["Random_Instances"] = df["account_no_norm"].apply(
            lambda x: x if x in sampled_accounts else ""
        )

        print(f"✔ Total approved sampled accounts (unique): {len(sampled_accounts)}")

        # ----------------------------
        # 🔍 Diagnostics / Verification
        # ----------------------------
        # Use safe boolean filtering
        sample_new = df[(df["is_new_account"] == True) & (df["Random_Instances"] != "")]
        sample_old = df[
            (df["is_new_account"] == False) & (df["Random_Instances"] != "")
        ]

        # New accounts check
        print("\n--- New Accounts Sampling Verification ---")
        if not sample_new.empty:
            if "category" in sample_new.columns:
                print("Category distribution:\n", sample_new.groupby("category").size())
            print("Total new sampled accounts:", len(sample_new))
            cols_new = ["is_new_account", "category", "account_no_norm"]
            cols_new = [c for c in cols_new if c in sample_new.columns]
            print(sample_new[cols_new].head())
        else:
            print("No new accounts sampled.")

        # Old accounts check
        print("\n--- Old Accounts Sampling Verification ---")
        print("High Turnover Accounts:")
        if not old_sample_high.empty:
            cols_high = ["TOTAL_ACTUAL_TO_NUM", "account_no_norm"]
            cols_high = [c for c in cols_high if c in old_sample_high.columns]
            print(old_sample_high[cols_high].head(2))
        else:
            print("No high turnover old accounts sampled.")

        print("Other Category Accounts:")
        if not old_sample_other.empty:
            cols_other = ["CUSTOMER_OCCUPATION", "account_no_norm"]
            cols_other = [c for c in cols_other if c in old_sample_other.columns]
            print(old_sample_other[cols_other].head(2))
        else:
            print("No other category old accounts sampled.")

        print(
            "Total old sampled accounts:",
            len(pd.concat([old_sample_high, old_sample_other])),
        )

        # Combined table for review (use actual sampled subsets)
        all_sampled = pd.concat([sample_new, old_sample_high, old_sample_other])
        cols_all = [
            "account_no_norm",
            "is_new_account",
            "category",
            "TOTAL_ACTUAL_TO_NUM",
            "CUSTOMER_OCCUPATION",
        ]
        cols_all = [c for c in cols_all if c in all_sampled.columns]

        # Summary line for clarity
        print(
            f"Summary → Exceptions: {len(exception_accounts)}, "
            f"Old: {len(df_old)}, New: {len(df_new)}, "
            f"Sampled: {len(sampled_accounts)}"
        )

        # Save globally for later use
        self.approved_sample_accounts = sampled_accounts
        self.filtered_df = df[df["Random_Instances"] != ""].copy()
        return self.filtered_df

    # ============================================================
    # APPLY APPROVED SAMPLING TO KYC POPULATION SHEETS
    # ============================================================
    def apply_approved_sampling_to_kyc(self):
        if not hasattr(self, "approved_sample_accounts"):
            print("⚠️ No approved sampled accounts found.")
            return

        normalized_samples = {
            self.normalize_account(acc) for acc in self.approved_sample_accounts if acc
        }

        def apply_to_df(df, account_col):
            if df is None or df.empty:
                return df
            df = df.copy()
            if account_col in df.columns:
                df[account_col] = df[account_col].apply(self.normalize_account)
                df["Random_Instances"] = df[account_col].apply(
                    lambda x: x if x in normalized_samples else ""
                )
                # move column to end
                cols = list(df.columns)
                cols.append(cols.pop(cols.index("Random_Instances")))
                df = df[cols]
                print(
                    f"✅ {df['Random_Instances'].ne('').sum()} accounts marked in '{account_col}'"
                )
            else:
                df["Random_Instances"] = ""
            return df

        self.kyc_new_final = apply_to_df(self.kyc_new_final, "accountno")
        self.kyc_ht_final = apply_to_df(self.kyc_ht_final, "Account No.")
        self.kyc_ind_final = apply_to_df(self.kyc_ind_final, "Account_numumber")

        print("✔ Approved Sampling applied to KYC population sheets.")

    # ─────────────────────────────────────────────────────────
    # File Upload and Merge Handler
    # ─────────────────────────────────────────────────────────
    def upload_files(self):
        file_paths = filedialog.askopenfilenames(
            filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv")]
        )
        if not file_paths:
            return

        self.files_data.clear()
        for path in file_paths:
            try:
                filename = get_filename(path)
                if path.lower().endswith(".csv"):
                    df = load_file(path)
                    df = clean_dataframe(df)
                    self.files_data[filename] = ({"Sheet1": df}, path)
                else:
                    xls = pd.ExcelFile(path)
                    sheets = {
                        sheet: clean_dataframe(
                            xls.parse(
                                sheet,
                                dtype=str,
                                keep_default_na=False,  # prevent N/A, NA from becoming NaN
                            )
                        )
                        for sheet in xls.sheet_names
                    }
                    self.files_data[filename] = (sheets, path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load {path}\n{e}")

        if not self.files_data:
            return

        self.status_label.config(text="Status: Files loaded", foreground="#000000")

        # --- Update Branch label automatically from file with BranchCode + Branch_Name ---
        branch_info = None
        for filename, (sheets_dict, path) in self.files_data.items():
            for sheet_name, df in sheets_dict.items():
                if "BranchCode" in df.columns and "Branch_Name" in df.columns:
                    branch_code = str(df["BranchCode"].iloc[0])
                    branch_name = str(df["Branch_Name"].iloc[0])
                    branch_info = f"{branch_code} - {branch_name}"
                    break
            if branch_info:
                break

        if branch_info:
            self.branch_label.config(text=f"Branch: {branch_info}")
        else:
            self.branch_label.config(text="Branch: [Not determined yet]")

        # Determine output paths dynamically
        first_file_path = next(iter(self.files_data.values()))[1]
        base_path = os.path.dirname(first_file_path)
        xlsx_out_path = os.path.join(base_path, "merged_file.xlsx")

        # If merged file already exists → silently skip
        if os.path.exists(xlsx_out_path):
            return

        # Run merge
        try:
            merged_df = process_uploaded_files(
                self.files_data, save_filename="merged_file.xlsx"
            )
            if merged_df is not None and not merged_df.empty:
                # Save merged file path
                merged_file_path = os.path.join(base_path, "merged_file.xlsx")
                # Add merged DataFrame into uploaded files dictionary
                merged_filename = "merged_file.xlsx"
                self.files_data[merged_filename] = (
                    {"Merged": merged_df},
                    merged_file_path,
                )
        except Exception as e:
            messagebox.showerror("Merge Error", f"Failed to merge files\n{e}")

    # ─────────────────────────────────────────────────────────
    # Normalize DataFrames
    # ─────────────────────────────────────────────────────────
    def get_normalized_dataframes(self):
        """Returns a normalized dictionary of all uploaded sheets with cleaned column names.
        Keys are in the format: filename::sheetname"""
        dataframes = {}
        for filename, (sheets_dict, path) in self.files_data.items():
            for sheet_name, df in sheets_dict.items():
                key = f"{filename}::{sheet_name}"
                df.columns = df.columns.str.strip()
                dataframes[key] = df
        return dataframes

    # ─────────────────────────────────────────────────────────
    # Populate Treeview with DataFrame
    # ─────────────────────────────────────────────────────────
    def populate_tree(self, df: pd.DataFrame):
        if self.detail_tree is None:
            return

        # Clear existing data
        self.detail_tree.delete(*self.detail_tree.get_children())

        if df is None or df.empty:
            self.detail_tree["columns"] = []
            return

        cols = list(df.columns)
        self.detail_tree["columns"] = cols
        self.detail_tree["show"] = "headings"

        for col in cols:
            self.detail_tree.heading(col, text=col)
            self.detail_tree.column(col, width=120, anchor="w")

        for _, row in df.iterrows():
            values = ["" if pd.isna(v) else str(v) for v in row]
            self.detail_tree.insert("", "end", values=values)

    # ─────────────────────────────────────────────────────────
    # Update Summary Label with Row and Column Info
    # ─────────────────────────────────────────────────────────
    def update_summary(self):
        if self.filtered_df is not None:
            rows, cols = self.filtered_df.shape
            self.summary_label.config(
                text=f"Summary: {rows:,} rows × {cols:,} columns", foreground="#7f8c8d"
            )
        else:
            self.summary_label.config(text="Summary: N/A", foreground="#7f8c8d")

    def compute_distinct_accounts_for_sampling(self):
        # """Populate self.distinct_accounts_for_sampling from KYC-New Accounts."""
        if hasattr(self, "kyc_new_final") and not self.kyc_new_final.empty:
            self.distinct_accounts_for_sampling = set(
                self.kyc_new_final["accountno"].dropna().astype(str).unique()
            )
            print(
                f"✅ Distinct accounts computed for sampling: {len(self.distinct_accounts_for_sampling)}"
            )
        else:
            self.distinct_accounts_for_sampling = set()
            print("⚠️ KYC-New Accounts is empty; no distinct accounts found.")

    # ─────────────────────────────────────────────────────────
    # Generate Excel Report
    # ─────────────────────────────────────────────────────────
    def generate_excel_workbook_from_ui(self):
        self.report_mode = True
        from KYC_Viewer.main import generate_breach_table

        if not self.files_data:
            messagebox.showerror(
                "Error", "No files uploaded. Please upload files first."
            )
            return

        # 🔑 Reset state before generating a new report
        self.logic_outputs = {}
        self.kyc_new_final = None
        self.kyc_ht_final = None
        self.kyc_ind_final = None

        # 5) KYC processing helpers
        def get_col_ci(df, col_name):
            if df is None or df.empty:
                return pd.Series([], dtype="object")
            mapping = {c.upper(): c for c in df.columns}
            key = col_name.upper()
            return (
                df[mapping[key]]
                if key in mapping
                else pd.Series([""] * len(df), index=df.index)
            )

        def create_empty_kyc_df(expected_cols):
            return pd.DataFrame(columns=expected_cols)

        def enforce_columns_and_order(partial_df, expected_cols):
            out = pd.DataFrame(index=partial_df.index)
            for col in expected_cols:
                matches = [c for c in partial_df.columns if c.upper() == col.upper()]
                out[col] = partial_df[matches[0]].values if matches else ""
            return out.reset_index(drop=True)

        # Run all registered logic functions and capture results
        dataframes = self.get_normalized_dataframes()
        results = {}
        for logic_name in self.logic_names:
            df = generate_breach_table(
                logic_name, dataframes, report_mode=self.report_mode
            )
            results[logic_name] = df

        # ✅ Save results on the object for later use
        self.logic_outputs = {
            name: {"df": df, "metadata": {}} for name, df in results.items()
        }

        # 2) Ask for filename & path to save
        custom_name = self.prompt_for_filename()
        if not custom_name:
            return
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel Workbook As",
            initialfile=f"{custom_name}.xlsx",
        )
        if not output_path:
            return

        # ✅ Save path for later use
        self.generated_report_path = output_path

        # 3) Prepare workbook and common styles
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF", size=12)
        table_font = Font(size=11)
        red_meta_font = Font(bold=True, color="FF0000", size=11)
        normal_meta_font = Font(bold=True, color="000000", size=11)
        metadata_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        table_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        RED_KEYS = {
            "Exception Text",
            "Risk",
            "Control Number",
            "eAudit Head Name",
            "Instances",
            "Control Effectiveness",
            "Note",
        }

        # --- Helper: safe writer for any DataFrame ---
        def write_kyc_sheet(sheet_name, df):
            sheet = wb.create_sheet(title=sheet_name[:31])
            if df is None or not isinstance(df, pd.DataFrame):
                df = pd.DataFrame()
            if df.columns.size > 0:
                sheet.append(list(df.columns))
                for col_idx, _ in enumerate(df.columns, 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="BDD7EE", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            else:
                sheet.append(["No data available"])

            if not df.empty and df.columns.size > 0:
                for i, row in df.iterrows():
                    sheet.append([v if v is not None else "" for v in row])
                    r = sheet.max_row
                    stripe_fill = PatternFill(
                        start_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                        fill_type="solid",
                    )
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = sheet.cell(row=r, column=col_idx)
                        cell.font = table_font
                        cell.fill = stripe_fill
                        cell.alignment = table_alignment
                        cell.border = thin_border

            try:
                for col in sheet.columns:
                    max_len = max(
                        (len(str(c.value)) if c.value else 0 for c in col), default=0
                    )
                    sheet.column_dimensions[get_column_letter(col[0].column)].width = (
                        max_len + 5
                    )
            except Exception:
                pass
            sheet.freeze_panes = "A2" if df.columns.size > 0 else "A1"

        # ----------------------------------------------------------
        # Apply sampling AFTER KYC DataFrames are finalized
        # ----------------------------------------------------------
        if getattr(self, "selected_sampling_type", None) == "Program Based Instances":
            self.distinct_accounts_for_sampling = self.extract_distinct_accounts()
            self.apply_program_sampling_to_kyc(self.distinct_accounts_for_sampling)
            print("✔ Program Sampling applied successfully.")

        elif getattr(self, "selected_sampling_type", None) == "Random Instances":
            print("===== APPROVED SAMPLING =====")

            # Build population DataFrame using helper (merges KYC + turnover)
            df_population = self.build_sampling_population(self.files_data)

            # Run approved sampling on the merged population
            df_sampled = self.approved_sampling(df_population)

            # Save for UI display: only the 109 sampled accounts
            sampled_rows = []
            for df_attr in ["kyc_new_final", "kyc_ht_final", "kyc_ind_final"]:
                df = getattr(self, df_attr, None)
                if df is not None and "Random_Instances" in df.columns:
                    subset = df[df["Random_Instances"] != ""].copy()
                    if not subset.empty:
                        sampled_rows.append(subset)

            if sampled_rows:
                self.filtered_df = pd.concat(sampled_rows, ignore_index=True)
            else:
                self.filtered_df = pd.DataFrame()  # fallback if nothing matched

            # Update UI grid and summary
            self.populate_tree(self.filtered_df)
            self.update_summary()

        # 6) Identify KYC file
        signature_set = {"ACCT_OPR_INST_1", "ID_VAL_DATE", "ID_NUMBER"}
        kyc_file_df = None
        for filename, (sheets_dict, path) in self.files_data.items():
            for sheet_name, df in sheets_dict.items():
                if signature_set.issubset({c.upper() for c in df.columns}):
                    kyc_file_df = df.copy()
                    break
            if kyc_file_df is not None:
                break

        self.kyc_new_final = create_empty_kyc_df(KYC_NEW_ACCOUNTS_COLS)
        self.kyc_ht_final = create_empty_kyc_df(KYC_OLD_ACCOUNTS_HT_COLS)
        self.kyc_ind_final = create_empty_kyc_df(KYC_OLD_ACCOUNTS_INDIV_COLS)

        if kyc_file_df is not None:
            kyc_file_df.columns = kyc_file_df.columns.str.strip()
            status_series = (
                get_col_ci(kyc_file_df, "StatusCode").astype(str).str.strip()
            )
            mask_current = status_series.str.upper() == "CURRENT"
            kyc_file_filtered = kyc_file_df[mask_current].copy()

            # Build partial KYC DataFrames
            kyc_new_partial = pd.DataFrame(
                {
                    "Serial_No": get_col_ci(kyc_file_filtered, "Account_Num"),
                    "branch_code": get_col_ci(kyc_file_filtered, "BranchCode"),
                    "customer_id": get_col_ci(kyc_file_filtered, "Customer_Num"),
                    "accountno": get_col_ci(kyc_file_filtered, "Account_Num"),
                    "account_type_(new)": get_col_ci(
                        kyc_file_filtered, "Cust_Sector_Code"
                    ),
                    "account_title": get_col_ci(kyc_file_filtered, "Title_of_Account"),
                    "currency": get_col_ci(kyc_file_filtered, "Currency"),
                    "category": get_col_ci(kyc_file_filtered, "Product_Code"),
                    "actual_currency": get_col_ci(kyc_file_filtered, "Currency"),
                    "ac_open_date": get_col_ci(kyc_file_filtered, "Account_Open_Date"),
                    "posting_rest": get_col_ci(kyc_file_filtered, "StatusCode"),
                }
            )
            kyc_ht_partial = pd.DataFrame(
                {
                    "Branch": get_col_ci(kyc_file_filtered, "BranchCode"),
                    "Br. Code & Name": get_col_ci(kyc_file_filtered, "Account_Num"),
                    "Customer ID": get_col_ci(kyc_file_filtered, "Customer_Num"),
                    "Account No.": get_col_ci(kyc_file_filtered, "Account_Num"),
                    "Account Title": get_col_ci(kyc_file_filtered, "Title_of_Account"),
                    "Account Type": get_col_ci(kyc_file_filtered, "Product_Code"),
                    "Entity Type (Sector)": get_col_ci(
                        kyc_file_filtered, "Sector_Description"
                    ),
                    "Currency": get_col_ci(kyc_file_filtered, "Currency"),
                    "Purpose of Account": get_col_ci(kyc_file_filtered, "Purpose"),
                    "Occupation": get_col_ci(kyc_file_filtered, "Cust_Occupation"),
                    "Nature of Business": get_col_ci(
                        kyc_file_filtered, "Nature_of_Business"
                    ),
                    "Account T/O (P.A KYC)": get_col_ci(
                        kyc_file_filtered, "KYC_Annual_Turnover"
                    ),
                    "ID No. of Customer": get_col_ci(kyc_file_filtered, "ID_Number"),
                    "Company": get_col_ci(kyc_file_filtered, "StatusCode"),
                }
            )
            kyc_ind_partial = pd.DataFrame(
                {
                    "Account_numumber": get_col_ci(kyc_file_filtered, "Account_Num"),
                    "Customer_Num": get_col_ci(kyc_file_filtered, "Customer_Num"),
                    "Name": get_col_ci(kyc_file_filtered, "Title_of_Account"),
                    "Account_Open_Dt": get_col_ci(
                        kyc_file_filtered, "Account_Open_Date"
                    ),
                    "Branch_Code": get_col_ci(kyc_file_filtered, "BranchCode"),
                    "Branch_name": get_col_ci(kyc_file_filtered, "Branch_Name"),
                    "EXP_MONTH_TOVER": get_col_ci(kyc_file_filtered, "MONTH_TOVER_RG"),
                    "KYC_ATO": get_col_ci(kyc_file_filtered, "KYC_Annual_Turnover"),
                    "MODEDEPOSITS": get_col_ci(kyc_file_filtered, "MODEDEPOSITS"),
                    "MODEWITHDRAW": get_col_ci(kyc_file_filtered, "MODEWITHDRAW"),
                    "Purpose": get_col_ci(kyc_file_filtered, "Purpose"),
                    "UNSCLISTST": get_col_ci(kyc_file_filtered, "UNSCLISTST"),
                    "CURRENCY": get_col_ci(kyc_file_filtered, "Currency"),
                    "CATEGORY": get_col_ci(kyc_file_filtered, "Product_Code"),
                    "Product_Desc": get_col_ci(kyc_file_filtered, "ProductDesc"),
                    "Status_of_Account": get_col_ci(
                        kyc_file_filtered, "StatusCode"
                    ).replace(
                        {"Current": "Active", "current": "Active", "CURRENT": "Active"}
                    ),
                    "Risk_Level": get_col_ci(kyc_file_filtered, "KYC_Risk"),
                    "SOURCE_OF_INCOME": get_col_ci(
                        kyc_file_filtered, "SOURCE_OF_INCOME"
                    ),
                    "OCCUPATION": get_col_ci(kyc_file_filtered, "Cust_Occupation"),
                    "NAME_OF_EMP": get_col_ci(kyc_file_filtered, "Name_of_Employer"),
                    "CURRENT_SALARY": get_col_ci(
                        kyc_file_filtered, "Cust_Current_Salary"
                    ),
                    "NAME_OF_BUS": get_col_ci(kyc_file_filtered, "Name_of_Business"),
                    "KYC_REVW_COMENT": get_col_ci(
                        kyc_file_filtered, "StatusCode"
                    ).replace(
                        {"Current": "Active", "current": "Active", "CURRENT": "Active"}
                    ),
                }
            )

            # Enforce final shapes
            self.kyc_new_final = enforce_columns_and_order(
                kyc_new_partial, KYC_NEW_ACCOUNTS_COLS
            )
            self.kyc_ht_final = enforce_columns_and_order(
                kyc_ht_partial, KYC_OLD_ACCOUNTS_HT_COLS
            )
            self.kyc_ind_final = enforce_columns_and_order(
                kyc_ind_partial, KYC_OLD_ACCOUNTS_INDIV_COLS
            )

            # ----------------------------------------------------------
            # Apply Approved Sampling AFTER DataFrames are rebuilt
            # ----------------------------------------------------------
            if getattr(self, "selected_sampling_type", None) == "Random Instances":
                self.apply_approved_sampling_to_kyc()
                print("✔ Approved_Sample column applied to KYC DataFrames.")

            # ----------------------------------------------------------------
            # >>>>> ACCOUNT AGE CLASSIFICATION (FIXED) <<<<<
            # ----------------------------------------------------------------
            if hasattr(self, "audit_start_date") and hasattr(self, "audit_end_date"):

                start_int = int(self.audit_start_date)
                end_int = int(self.audit_end_date)

                def date_to_int(val):
                    try:
                        return int(pd.to_datetime(str(val)).strftime("%Y%m%d"))
                    except:
                        return None  # safely handle invalid/missing dates

                if not self.kyc_new_final.empty:
                    kyc_dates = self.kyc_new_final["ac_open_date"].apply(date_to_int)

                    # safely mark new accounts
                    self.kyc_new_final["_is_new"] = kyc_dates.apply(
                        lambda x: start_int <= x <= end_int if x is not None else False
                    )

                # Compute totals using the temporary column
                self.total_new = self.kyc_new_final["_is_new"].sum()
                self.total_old = len(self.kyc_new_final) - self.total_new

                print("===== ACCOUNT AGE ANALYSIS =====")
                print("Total:", len(self.kyc_new_final))
                print("Old accounts:", self.total_old)
                print("New accounts:", self.total_new)

                # Drop the helper column before export so it doesn't appear in Excel
                if "_is_new" in self.kyc_new_final.columns:
                    self.kyc_new_final = self.kyc_new_final.drop(columns=["_is_new"])

            # -----------------------------------------------------------
            # ACCOUNT AGE CLASSIFICATION (NO FILTERING APPLIED)
            # -----------------------------------------------------------
            def add_audit_period(self):
                audit_start = self.start_entry.get().strip()
                audit_end = self.end_entry.get().strip()

                if not audit_start or not audit_end:
                    messagebox.showerror(
                        "Error", "Please enter both start and end dates."
                    )
                    return

                try:
                    self.audit_start = int(audit_start)
                    self.audit_end = int(audit_end)
                    messagebox.showinfo("Success", "Audit period saved successfully.")
                except:
                    messagebox.showerror("Error", "Invalid date format.")

            # Compute distinct accounts for sampling if not already present
            if (
                not hasattr(self, "distinct_accounts_for_sampling")
                or not self.distinct_accounts_for_sampling
            ):
                # if you have compute_distinct_accounts_for_sampling defined, call it
                try:
                    self.compute_distinct_accounts_for_sampling()
                except Exception:
                    # fallback: compute from kyc_new_final (accountno column)
                    try:
                        self.distinct_accounts_for_sampling = set(
                            self.kyc_new_final["accountno"]
                            .dropna()
                            .astype(str)
                            .str.strip()
                            .unique()
                        )
                        print(
                            f"Fallback distinct accounts computed: {len(self.distinct_accounts_for_sampling)}"
                        )
                    except Exception:
                        self.distinct_accounts_for_sampling = set()
                        print(
                            "Failed to compute distinct accounts for sampling (fallback)."
                        )

            # ----------------------------------------------------------
            # PROGRAM SAMPLING (FIXED)
            # ----------------------------------------------------------
            if self.selected_option.get() == "Program Based Instances":
                print("\n===== PROGRAM SAMPLING =====")

                # Collect exception accounts from extractor instead of logic_outputs
                exception_accounts_set = self.extract_distinct_accounts()

                print(f"✔ Total exception accounts: {len(exception_accounts_set)}")

                # ---- FIX: Identify TRUE account column for each KYC sheet ----
                ACCOUNT_COL_MAP = {
                    "kyc_new_final": "accountno",
                    "kyc_ht_final": "Account No.",
                    "kyc_ind_final": "Account_numumber",  # check typo here
                }

                for df_attr, acc_col in ACCOUNT_COL_MAP.items():
                    df = getattr(self, df_attr)

                    if df is None or df.empty:
                        continue

                    # Ensure column exists
                    if acc_col not in df.columns:
                        df["Program_Based _Instances"] = ""
                        setattr(self, df_attr, df)
                        continue

                    # Normalize exception accounts
                    exception_normalized = {
                        self.normalize_account(x) for x in exception_accounts_set if x
                    }

                    # Normalize population accounts and mark matches
                    df[acc_col] = df[acc_col].apply(self.normalize_account)
                    df["Program_Based _Instances"] = df[acc_col].apply(
                        lambda x: x if x in exception_normalized else ""
                    )

                    matched = (df["Program_Based _Instances"] != "").sum()
                    print(
                        f"✅ {matched} accounts matched in sheet using column '{acc_col}'"
                    )

                    setattr(self, df_attr, df)

                print("✔ Program Sampling applied successfully.")

            # ============================================================================
            # CONTROL EFFECTIVENESS (RUN AFTER SAMPLING)
            # ============================================================================
            if self.selected_option.get() in [
                "Program Based Instances",
                "Random Instances",
            ]:

                sample_col = (
                    "Program_Based _Instances"
                    if self.selected_option.get() == "Program Based Instances"
                    else "Random_Instances"
                )

                control_effectiveness_map = {}

                def normalize_logic_name(name):
                    return (
                        str(name).strip().replace(" ", "_").replace("__", "_").upper()
                    )

                def normalize_acc(x):
                    if pd.isna(x) or x == "":
                        return None
                    x = str(x).strip().upper()
                    x = re.sub(r"[^A-Z0-9]", "", x)
                    return x if x else None

                # Gather all sample accounts from population sheets once
                population_accounts = set()
                for df_attr in ["kyc_new_final", "kyc_ht_final", "kyc_ind_final"]:
                    kyc_df = getattr(self, df_attr, pd.DataFrame())
                    if sample_col in kyc_df.columns:
                        population_accounts.update(
                            filter(None, (normalize_acc(x) for x in kyc_df[sample_col]))
                        )

                for logic_name, info in self.logic_outputs.items():

                    df = info.get("df")
                    effectiveness_str = "Not Calculated"
                    num_instances = 0

                    # Skip if df is invalid
                    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                        if logic_name in logic_metadata:
                            logic_metadata[logic_name][
                                "Control Effectiveness"
                            ] = effectiveness_str
                            logic_metadata[logic_name]["Instances"] = num_instances
                        continue

                    # Safely get exception accounts
                    if "Account_Number" in df.columns:
                        exception_accts = set(
                            filter(
                                None, (normalize_acc(x) for x in df["Account_Number"])
                            )
                        )
                    else:
                        exception_accts = set()
                    # Count only rows with a non‑empty Account_Number
                    # Normalize column names to uppercase for consistency
                    df.columns = [str(c).strip().upper() for c in df.columns]

                    # Count only rows with a non-empty ACCOUNT_NUMBER
                    if "ACCOUNT_NUMBER" in df.columns and not df.empty:
                        num_instances = (
                            df["ACCOUNT_NUMBER"]
                            .dropna()
                            .astype(str)
                            .str.strip()
                            .ne("")
                            .sum()
                        )
                    else:
                        num_instances = 0

                    # CE calculation
                    if self.selected_option.get() == "Program Based Instances":
                        # Correct formula: total exception accounts / total program sample accounts
                        total_program_accounts = len(population_accounts)
                        ce_percent = (
                            (len(exception_accts) / total_program_accounts * 100)
                            if total_program_accounts
                            else 0
                        )
                    else:
                        # Approved Sampling CE: matched / total approved sample accounts
                        matched = exception_accts & population_accounts
                        total_sample_count = len(population_accounts)
                        ce_percent = (
                            (len(matched) / total_sample_count * 100)
                            if total_sample_count
                            else 0
                        )

                    # Assign label based on thresholds
                    if ce_percent <= 25:
                        label = "Control is Effective"
                    elif ce_percent <= 50:
                        label = "Control Effective With Some Exceptions"
                    else:
                        label = "Control Not Effective"

                    effectiveness_str = label

                    # Update metadata
                    info.setdefault("metadata", {})
                    info["metadata"]["Control Effectiveness"] = effectiveness_str
                    info["metadata"]["Instances"] = num_instances

                    if logic_name in logic_metadata:
                        logic_metadata[logic_name][
                            "Control Effectiveness"
                        ] = effectiveness_str
                        logic_metadata[logic_name]["Instances"] = num_instances

                    control_effectiveness_map[normalize_logic_name(logic_name)] = (
                        effectiveness_str
                    )

                # Merge into Exception Summary sheet
                if hasattr(self, "exception_summary_df") and isinstance(
                    self.exception_summary_df, pd.DataFrame
                ):
                    self.exception_summary_df["Logic_Name_clean"] = (
                        self.exception_summary_df["Logic_Name"].apply(
                            normalize_logic_name
                        )
                    )
                    ce_df = pd.DataFrame(
                        [(k, v) for k, v in control_effectiveness_map.items()],
                        columns=["Logic_Name_clean", "Control Effectiveness"],
                    )
                    self.exception_summary_df = self.exception_summary_df.merge(
                        ce_df, on="Logic_Name_clean", how="left"
                    )
                    self.exception_summary_df.drop(
                        columns=["Logic_Name_clean"], inplace=True
                    )
                    self.exception_summary_df["Control Effectiveness"] = (
                        self.exception_summary_df["Control Effectiveness"].fillna(
                            "Not Calculated"
                        )
                    )

                    # Write Exception Summary sheet
                    write_kyc_sheet("Exception Summary", self.exception_summary_df)

                print("✔ Control Effectiveness calculated and written successfully.")

            # 4) Write exception logic sheets
            for logic_name, info in self.logic_outputs.items():

                # Use result DF
                df = info.get("df")
                if df is None or not isinstance(df, pd.DataFrame):
                    df = pd.DataFrame()
                df = df.replace({pd.NA: None}).applymap(
                    lambda x: None if pd.isna(x) or str(x).strip() == "" else x
                )
                df = df.dropna(how="all").copy()

                # -----------------------------------------
                # ALWAYS fetch fresh metadata updated by CE
                # -----------------------------------------
                updated_meta = logic_metadata.get(logic_name, {}).copy()

                # Pull CE + Instances from info['metadata'] if available
                meta = info.get("metadata", {})

                # Control Effectiveness
                if "Control Effectiveness" in meta:
                    ce_val = meta["Control Effectiveness"]
                elif "Control_Effectiveness" in df.columns:
                    ce_val = df["Control_Effectiveness"].iloc[0]
                else:
                    ce_val = "Not Calculated"

                # Always set Instances to the number of rows in df
                inst_val = len(df) if isinstance(df, pd.DataFrame) else 0

                # Update the metadata
                updated_meta["Control Effectiveness"] = ce_val
                updated_meta["Instances"] = inst_val

                # Create sheet
                sheet = wb.create_sheet(title=logic_name[:31])

                # --- Write Metadata ---
                for key, value in updated_meta.items():
                    sheet.append([key, value])
                    row = sheet.max_row
                    key_cell = sheet[f"A{row}"]
                    val_cell = sheet[f"B{row}"]

                    # Color code
                    key_cell.font = (
                        red_meta_font if key in RED_KEYS else normal_meta_font
                    )
                    key_cell.fill = metadata_fill
                    val_cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")

                    # Borders & alignment
                    key_cell.border = thin_border
                    val_cell.border = thin_border
                    key_cell.alignment = table_alignment
                    val_cell.alignment = table_alignment

                sheet.append([])

                # --- Write Table Header ---
                if df.columns.size > 0:
                    sheet.append(list(df.columns))
                    header_row = sheet.max_row
                    for col_idx, _ in enumerate(df.columns, 1):
                        cell = sheet.cell(row=header_row, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = table_alignment
                        cell.border = thin_border

                # --- Write Table Rows ---
                if not df.empty and df.columns.size > 0:
                    for i, row in df.iterrows():
                        sheet.append([v if v is not None else "" for v in row])
                        r = sheet.max_row
                        stripe_fill = PatternFill(
                            start_color="F2F2F2" if i % 2 == 0 else "FFFFFF",
                            fill_type="solid",
                        )
                        for col_idx in range(1, len(df.columns) + 1):
                            cell = sheet.cell(row=r, column=col_idx)
                            cell.font = table_font
                            cell.fill = stripe_fill
                            cell.alignment = table_alignment
                            cell.border = thin_border

                # Auto column width
                try:
                    for col in sheet.columns:
                        max_len = max(
                            (len(str(c.value)) if c.value else 0 for c in col),
                            default=0,
                        )
                        sheet.column_dimensions[
                            get_column_letter(col[0].column)
                        ].width = (max_len + 5)
                except:
                    pass

                sheet.freeze_panes = "A3"

            # ----------------------------------------------------------
            # Write KYC sheets ONCE (outside the loop)
            # ----------------------------------------------------------
            write_kyc_sheet("KYC-New Accounts", self.kyc_new_final)
            write_kyc_sheet("KYC-Old Accounts - High Turnover", self.kyc_ht_final)
            write_kyc_sheet("KYC-Old Accounts - Individual", self.kyc_ind_final)

            # ----------------------------------------------------------
            # Save workbook ONCE (outside the loop)
            # ----------------------------------------------------------
            try:
                wb.save(output_path)
                messagebox.showinfo(
                    "Success", f"Excel workbook saved to:\n{output_path}"
                )
            except Exception as e:
                messagebox.showerror("Save Error", f"Failed to save workbook:\n{e}")

            # Enable Add Audit Period button
            if hasattr(self, "audit_btn"):
                self.audit_btn.config(state="normal")

    def get_branch_name_from_uploaded(self):
        """Return branch name from the merged file (first non-empty Branch_Code)."""
        for filename, (sheets_dict, path) in self.files_data.items():
            # Check for merged file first
            if "Merged" in sheets_dict:
                df = sheets_dict["Merged"]
                if "Branch_Code" in df.columns:
                    # Take the first non-null value
                    branch_code = df["Branch_Code"].dropna().iloc[0]
                    return str(branch_code)
        # fallback
        return "[Not determined yet]"

    # ─────────────────────────────────────────────────────────
    # PDF / Excel Filename Prompt Dialog
    # ─────────────────────────────────────────────────────────
    def prompt_for_filename(self):

        # 🪟 Create modal dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Enter Filename")
        dialog.geometry("400x150")
        dialog.resizable(False, False)
        dialog.grab_set()

        # 🎯 Center the dialog on screen
        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        x = (dialog.winfo_screenwidth() // 2) - (w // 2)
        y = (dialog.winfo_screenheight() // 2) - (h // 2)
        dialog.geometry(f"{w}x{h}+{x}+{y}")

        # 📝 Input field and label
        tk.Label(
            dialog,
            text="Enter desired filename (without .xlsx):",
            font=("Segoe UI", 10),
        ).pack(pady=(20, 5))

        entry_var = tk.StringVar(value="KYC_Report")
        entry = tk.Entry(
            dialog, textvariable=entry_var, font=("Segoe UI", 10), width=40
        )
        entry.pack(pady=5)
        entry.focus()

        result = {"filename": None}

        # ✅ Submit logic with validation
        def on_submit():
            name = entry_var.get().strip()
            if name and re.match(r"^[\w\- ]+$", name):
                result["filename"] = name
                dialog.destroy()
            else:
                messagebox.showerror(
                    "Invalid Name",
                    "Please enter a valid filename without special characters.",
                )

        # ❌ Cancel logic
        def on_cancel():
            dialog.destroy()

        # 🎛 Button layout
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="OK", command=on_submit, width=10).pack(
            side="left", padx=10
        )
        tk.Button(btn_frame, text="Cancel", command=on_cancel, width=10).pack(
            side="right", padx=10
        )

        # ⏳ Wait for user input
        self.root.wait_window(dialog)
        return result["filename"]

    def run_logic_and_show(self, logic_name: str):
        if not self.files_data:
            messagebox.showwarning("Missing Files", "Please upload at least one file.")
            return

        logic_entry = logic_registry.get(logic_name)
        if not logic_entry or "function" not in logic_entry:
            messagebox.showerror("Error", f"Logic '{logic_name}' not found.")
            return

        logic_func = logic_entry["function"]

        # Build normalized dataframes
        dataframes = self.get_normalized_dataframes()

        try:
            result_df = logic_func(dataframes, mode="ui")
        except Exception as e:
            messagebox.showerror(
                "Error", f"Failed to run logic '{logic_name}':\n{type(e).__name__}: {e}"
            )
            return

        if result_df is None or result_df.empty:
            self.filtered_df = pd.DataFrame()
            self.populate_tree(self.filtered_df)
            self.summary_label.config(
                text="Summary: 0 rows × 0 columns", foreground="#7f8c8d"
            )
        else:
            self.filtered_df = result_df
            self.populate_tree(result_df)
            self.update_summary()

    # Map observation number → logic name (simple mapping for now)
    def get_logic_for_observation(self, obs_number: int):
        idx = obs_number - 1
        if 0 <= idx < len(self.logic_names):
            return self.logic_names[idx]
        return None

    # ─────────────────────────────────────────────────────────
    # Run Selected Logic and Display Table
    # ─────────────────────────────────────────────────────────
    def run_selected_logic(self):
        logic_name = self.selected_logic.get()
        if not logic_name:
            messagebox.showwarning("Missing Logic", "Please select a logic function.")
            return

        logic_entry = logic_registry.get(logic_name)
        if not logic_entry or "function" not in logic_entry:
            messagebox.showerror("Error", f"Logic '{logic_name}' not found or invalid.")
            return

        logic_func = logic_entry["function"]

        # Build full dataframes dict from uploaded files
        files_data = {}
        for filename, (sheets_dict, path) in self.files_data.items():
            for sheet_name, df in sheets_dict.items():
                key = f"{filename}::{sheet_name}"
                df.columns = df.columns.str.strip()
                files_data[key] = df

        # Run logic
        try:
            result_df = logic_func(files_data.copy(), mode="full")
            if result_df is not None and not result_df.empty:
                self.filtered_df = result_df
                self.populate_tree(result_df)
                self.update_summary()
                self.status_label.config(
                    text=f"Status: Table generated using '{logic_name}'",
                    foreground="#27ae60",
                )
            else:
                messagebox.showinfo("No Data", "Logic returned empty table.")
        except Exception as e:
            messagebox.showerror(
                "Error", f"Failed to run logic '{logic_name}': {type(e).__name__}: {e}"
            )

    # ─────────────────────────────────────────────────────────
    # Observation grid (left side)
    # ─────────────────────────────────────────────────────────
    def create_observation_grid(self, parent):
        max_obs = 107
        cols = 3

        for c in range(cols):
            parent.columnconfigure(c, weight=1)

        for obs in range(1, max_obs + 1):
            r = (obs - 1) // cols
            c = (obs - 1) % cols

            logic_name = self.get_logic_for_observation(
                obs
            )  # get the actual registered name

            btn = tk.Button(
                parent,
                text=str(obs),
                font=("Arial Black", 10, "bold"),
                bg="#DA1E26",
                fg="white",
                relief="raised",
                command=lambda ln=logic_name, n=obs: (
                    self.selected_logic.set(ln),  # use actual name
                    self.on_observation_clicked(n),
                ),
            )

            btn.grid(row=r, column=c, padx=2, pady=2, sticky="nsew")
            self.obs_buttons[obs] = btn

            # ------------------- Add tooltip -------------------
            # Get Exception Text from logic_metadata
            exception_text = logic_metadata.get(logic_name, {}).get(
                "Exception Text", "No details available"
            )
            Tooltip(btn, text=exception_text)

        rows = (max_obs + cols - 1) // cols
        for r in range(rows):
            parent.rowconfigure(r, weight=1)

    def on_observation_clicked(self, obs_number: int):
        self.selected_observation.set(obs_number)

        logic_name = self.get_logic_for_observation(obs_number)

        if logic_name:
            self.detail_title_label.config(
                text=f"Observation {obs_number}: {logic_name}"
            )
            # If files already uploaded, run and display
            if self.files_data:
                self.run_logic_and_show(logic_name)
        else:
            self.detail_title_label.config(
                text=f"Observation {obs_number}: [no logic linked]"
            )
            self.populate_tree(pd.DataFrame())  # clear table

        self.status_label.config(
            text=f"Status: Observation {obs_number} selected", foreground="#2980b9"
        )


# ─────────────────────────────────────────────────────────────
# Launch the App
# ─────────────────────────────────────────────────────────────
def render_ui(icon_path=None):
    root = tk.Tk()
    if icon_path:
        try:
            root.iconbitmap(
                r"F:\Wajid Hussain\Tasks\KYC_Model\KYC_Viewer\assets\BAFL.PK.ico"
            )
        except Exception as e:
            print(f"⚠️ Failed to load icon: {icon_path}")
            print(f"🧪 Exception: {e}")
    app = ExcelCSVViewer(root)
    root.mainloop()
